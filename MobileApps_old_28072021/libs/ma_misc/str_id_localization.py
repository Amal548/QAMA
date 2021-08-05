import os
import re
import json
import codecs
import difflib
import openpyxl
from datetime import datetime
import xml.etree.ElementTree as ET
from SAF.misc import package_utils
from SAF.misc.couch_wrapper import CouchWrapper
from MobileApps.libs.app_package.app_class_factory import app_module_factory

from io import open
try:
    import html
    from html.parser import HTMLParser
except ImportError:
    from HTMLParser import HTMLParser
    html = HTMLParser()

class MyHTMLParser(HTMLParser):
    def __init__(self):
        HTMLParser.__init__(self)
        self.recording = 0
        self.data = []

    def handle_data(self, data):
        self.data.append(data)

class StringProcessor(object):

    lang_list = ["en", "cs", "da", "de", "el", "es", "fi", "fr", "hu", "it", "ja", "ko", "nb", "nl", "pl", "ru", "sv", "tr", "bg", "et", "hr", "lt", "lv", "ro", "sk", "sl"]
    android_ch = ["ar", "iw", "pt", "zh-rCN", "zh-rTW"]
    ios_ch = ["pt-BR", "zh-Hans", "zh-Hant"]


    def __init__(self, os_type, db_connection):
        self.os_type = os_type
        self.db_connection = db_connection
        if self.os_type == "android":
            self.lang_list += self.android_ch
        elif self.os_type == "ios":
            self.lang_list += self.ios_ch


    def get_new_rec_id(self, project_name):
        result = self.db_connection.getViewResultByKey("get_info", "max_id", startKey=[self.os_type, project_name], endKey=[self.os_type, project_name], reduce=True)
        if len(result.rows) == 1:
            return int(result.rows[0]["value"]) + 1
        elif len(result.rows) == 0:
            return 0

    def pull_previous_records(self):
        pass

    def upload_record_to_couchdb(self, project_name, spec_version, pkg_source, apk_app_string_dict, spec_string_dict, result_dict, str_id_total_diff=None):
        rec= {
            "spec_version": spec_version,
            "project_name": project_name,
            "pkg_source": pkg_source,
            "os_type": self.os_type,
            "reviewed": False,
            "upload_time": datetime.now().strftime("%m/%d/%Y, %H:%M:%S"),
            "rec_id": str(self.get_new_rec_id(project_name)),
            "apk_str_dict": apk_app_string_dict, 
            "spec_str_dict": spec_string_dict, 
            "result_str_dict": result_dict,
            "str_id_total_diff": str_id_total_diff
        }
        self.db_connection.saveRecord(rec)

    def new_run(self, project_name, pkg_location, spec_location, test_run=True, include_previous=True):
        if self.os_type=="android":
            pkg_app_string_dict = self.rip_all_str_id_from_pkg(pkg_location)
            spec_string_dict = self.android_build_dict_from_xml(spec_location)

        elif self.os_type == "ios":
            pkg_app_string_dict = self.rip_all_str_id_from_pkg(pkg_location)
            spec_string_dict = self.ios_build_dict_from_xliff(spec_location)

        elif self.os_type == "other":
            pkg_app_string_dict = self.medillia_build_dict_form_json_source(pkg_location)
            spec_string_dict = self.medillia_build_dict_from_xlsx(spec_location)


        results_dict = self.run_results(project_name, pkg_app_string_dict, spec_string_dict, include_previous=include_previous)
        if include_previous:
            prev_record = self.load_prev_results(project_name)
            if prev_record is not None:
                str_id_diff_total = len(pkg_app_string_dict.keys()) -  len(prev_record["apk_str_dict"].keys())
            else:
                str_id_diff_total = None
        else:
            str_id_diff_total = None

        if not test_run:
            spec_version = spec_location.split("/")[-1]
            self.upload_record_to_couchdb(project_name, spec_version, pkg_location, pkg_app_string_dict, spec_string_dict, results_dict, str_id_total_diff=str_id_diff_total)
        else:
            return {"result_dict": results_dict, "spec_total": len(spec_string_dict.keys()), "str_id_total_diff": str_id_diff_total}


    def medillia_build_dict_from_xlsx(self, xlsx_path):
        data = {}
        book = openpyxl.load_workbook(xlsx_path)
        for item in book.sheetnames:
            sheet = book[item]
            id = 0
            language = None
            for column in range(sheet.max_column):
                id = 0
                for row in range(sheet.max_row): 
                    if row==0:
                        language = sheet.cell(row=row+1, column=column+1).value.split(" ")[-1].lower()
                        continue
                    elif row in [1, 2]:
                        continue
                    str_id = item+"_"+str(id)
                    if sheet.cell(row=row+1, column=column+1).value is None:
                        break
                    if data.get(str_id, None) is None:
                        data[str_id] = {language:sheet.cell(row=row+1, column=column+1).value.strip()}
                    else:
                        data[str_id][language] = sheet.cell(row=row+1, column=column+1).value.strip()
                    id += 1
        return data

    def medillia_build_dict_form_json_source(self, json_path):
        parser = MyHTMLParser()
        fh = open(json_path, "r", encoding="utf-8")
        data_file = json.load(fh)
        return_data = {}
        for _a in data_file["propertyConfiguration"]["forms"]:
            gi = False
            if "TEST" in _a["formJson"]["name"]:
                continue
            if "General Intercept" in _a["formJson"]["name"]:
                gi = True
                sheet_name = "General Intercept"
            else: 
                sheet_name = _a["formJson"]["name"].split(" ")[-2]
            id = 0
            language = _a["formJson"]["name"].split(" ")[-1].lower()
            got_placeholder = False
    
            actual_id = sheet_name+"_"+str(id)

            for _b in _a["formJson"]["pages"]:

                for _c in _b["dynamicData"]:
                    if _c["labelContent"] == "":
                        if return_data.get(actual_id, None) is None:
                            return_data[actual_id] = {language: _c["label"].strip()}
                        else:
                            return_data[actual_id][language] = _c["label"].strip()
                        id += 1
                        if id == 28:
                            id+=1
                        actual_id = sheet_name+"_"+str(id)
                    else:
                        parser.data=[]
                        parser.feed(html.unescape(_c["labelContent"]))
                        if return_data.get(actual_id, None) is None:
                            return_data[actual_id] = {language: "".join(parser.data).strip("\n").strip()}
                        else:
                            return_data[actual_id][language] = "".join(parser.data).strip("\n").strip()
                        id += 1   
                        if id == 28:
                            id+=1
                        actual_id = sheet_name+"_"+str(id)

                    if _c["optionsById"] is not None:
                        for d in _c["optionsById"]:
                            if return_data.get(actual_id, None) is None:
                                return_data[actual_id] = {language: d["label"].strip()}
                            else:
                                return_data[actual_id][language] = d["label"].strip()
                            id += 1 
                            if id == 28:
                                id+=1
                            actual_id = sheet_name+"_"+str(id)
                    if _c.get("ratingScales", None) is not None:
                        if return_data.get(actual_id, None) is None:
                            return_data[actual_id] = {language: _c["ratingScales"][0]["label"].strip()}
                        else:
                            return_data[actual_id][language] = _c["ratingScales"][0]["label"].strip()
                        id += 1 
                        if id == 28:
                            id+=1
                        actual_id = sheet_name+"_"+str(id)

                        if return_data.get(actual_id, None) is None:
                            return_data[actual_id] = {language: _c["ratingScales"][-1]["label"].strip()}
                        else:
                            return_data[actual_id][language] = _c["ratingScales"][-1]["label"].strip()
                        id += 1 
                        if id == 28:
                            id+=1
                        actual_id = sheet_name+"_"+str(id)

                    if not got_placeholder and _c.get("placeholder", "") not in ["", "Insert label text"]:

                        if return_data.get(actual_id, None) is None:
                            return_data[sheet_name+"_"+str(28)] = {language: _c["placeholder"].strip()}
                        else:
                            return_data[sheet_name+"_"+str(28)][language] = _c["placeholder"].strip()
                        got_placeholder = True

            id = 0
            if not gi and _a["formJson"]["settings"].get("formBasicSettings", {}).get("submitButtonLabel", "") != "":
                if return_data.get(actual_id, None) is None:
                    return_data[actual_id] = {language: _a["formJson"]["settings"]["formBasicSettings"]["submitButtonLabel"]}
                else:
                    return_data[actual_id][language] = _a["formJson"]["settings"]["formBasicSettings"]["submitButtonLabel"]
                id += 1 
                if id == 28:
                    id+=1
                actual_id = sheet_name+"_"+str(id)
            id = 0
            if gi:
                sheet_name += " Invite"
                actual_id = sheet_name+"_"+str(id)
                invite_field_keys = ["invitationHeadline","invitationText","provideButtonText","declineButtonText","laterButtonText"]
                for key in invite_field_keys:
                    if return_data.get(actual_id, None) is None:
                        return_data[actual_id] = {language: _a["inviteData"][key].strip()}
                    else:
                        return_data[actual_id][language] = _a["inviteData"][key].strip()
                    id += 1 
                    if id == 28:
                        id+=1
                    actual_id = sheet_name+"_"+str(id)
        return return_data        

    def rip_all_str_id_from_pkg(self, pkg_location):
        method_dict = {"android": package_utils.get_app_string_from_apk,
                       "ios": package_utils.get_app_string_from_ipa}

        pkg_app_string_dict = {}
        for lang in self.lang_list:
            localized_str = method_dict[self.os_type](pkg_location, lang, eng_base=False)
            language = lang.split("/")[-1]
            for key, value in localized_str.items():
                if pkg_app_string_dict.get(key, None) is None:
                    pkg_app_string_dict[key] = {language:value}
                else:
                    pkg_app_string_dict[key][language] = value
        return pkg_app_string_dict

    def ios_build_dict_from_xliff(self, resource_folder):
        spec_app_string_dict = {}
        lang_dir = [os.path.join(resource_folder, x) for x in os.listdir(resource_folder) if os.path.isfile(os.path.join(resource_folder, x))]
        for lang_file in lang_dir:
            language = lang_file.split("/")[-1].split(".")[0]
            reg = ET.parse(lang_file)
            xmlns = reg.getroot().tag.split("}")[0] + "}"
            re_strings = reg.findall(".//" + xmlns + "trans-unit")
            for item in re_strings:
                target = item.find(xmlns + "target")
                str_txt = codecs.escape_decode(bytes(target.text, "utf-8"))[0].decode("utf-8").strip()
                if spec_app_string_dict.get(item.get("id"), None) is None:
                    spec_app_string_dict[item.get("id")] = {language: str_txt}
                else:
                    spec_app_string_dict[item.get("id")][language] = str_txt
        return spec_app_string_dict

    def android_build_dict_from_xml(self, path):
        spec_app_string_dict = {}
        res_path = path + "/res"
        if not os.path.isdir(path):
            raise OSError("Path: " + path + " is not a directory")
        if not os.path.isdir(res_path):
            raise OSError("Path: " + path + " does not contain the res folder")

        lang_dir = [os.path.join(res_path,x) for x in os.listdir(res_path) if os.path.isdir(os.path.join(res_path, x))]
        for lang in lang_dir:
            language = "-".join(lang.split("/")[-1].split("-")[1:]) if "-" in lang else "en"
            all_files = [os.path.join(lang,x) for x in os.listdir(lang)]
            lang_files = []
            for files in all_files:
                files_parts = files.split("/")[-1].split(".")[0].split("_")
                if files_parts[0] == "strings":
                    lang_files.append(files)
                else:
                    os.remove(files)
            for xml_files in lang_files:
                tree = ET.parse(xml_files)
                root = tree.getroot()
                for str_id in root.findall('string'):
                    if spec_app_string_dict.get(str_id.get("name"), {}).get(language, False):
                        raise ValueError("Error: " + str_id["name"] + " duplicated")
                    raw_string = u"".join([x for x in str_id.itertext()]).rstrip()
                    unescaped_str = codecs.escape_decode(bytes(raw_string, "utf-8"))[0].decode("utf-8").strip()
                    if not spec_app_string_dict.get(str_id.get("name"), False):
                        spec_app_string_dict[str_id.get("name")] = {language: unescaped_str}
                    else:
                        spec_app_string_dict[str_id.get("name")][language] = unescaped_str
        return spec_app_string_dict

    def load_prev_results(self, project_name):
        prev_rec_id = self.get_new_rec_id(project_name) - 1
        db_rec = self.db_connection.getViewResultByKey("get_info", "rec_by_id", startKey=[self.os_type, project_name, str(prev_rec_id)], endKey=[self.os_type, project_name, str(prev_rec_id)])
        if len(db_rec.rows) != 1:
            #print "Parent record doesn't exist: " + str(prev_rec_id) + " " + self.os_type
            return None
        else:
            #print "Parent record found"
            return db_rec.rows[0]["value"]

    def run_results(self, project_name, apk_app_string_dict, spec_app_string_dict, include_previous=True):
        results_dict = {}
        if include_previous:
            prev_rec = self.load_prev_results(project_name)
            if prev_rec is not None:
                for str_id, info in prev_rec["result_str_dict"].items():
                    if info["result"] == "passed" or info["result"] == "removed" or info["result"] == "not test":
                        parent_spec_dict = prev_rec["spec_str_dict"].get(str_id, None)
                        parent_pkg_dict = prev_rec["apk_str_dict"].get(str_id, None)
                        cur_spec_dict = spec_app_string_dict.get(str_id, None)
                        cur_pkg_dict = apk_app_string_dict.get(str_id, None)

                        if parent_pkg_dict == cur_pkg_dict and parent_spec_dict == cur_spec_dict:
                            if parent_pkg_dict is None:
                                results_dict[str_id] = {"result": info["result"], "reason": "[3][" + info["result"] + "]Legacy: str_id not in apk", "reviewed": True}
                            elif parent_spec_dict is None:
                                results_dict[str_id] = {"result": info["result"], "reason": "[4][" + info["result"] + "]Legacy: str_id not in spec", "reviewed": True}
                            else:
                                results_dict[str_id] = {"result": info["result"], "reason": "[2][" + info["result"] + "]Legacy: str data unchanged", "reviewed": True}
                    elif info["result"] == "failed" and info.get("reviewed", False) is True:
                        parent_spec_dict = prev_rec["spec_str_dict"].get(str_id, None)
                        parent_pkg_dict = prev_rec["apk_str_dict"].get(str_id, None)
                        cur_spec_dict = spec_app_string_dict.get(str_id, None)
                        cur_pkg_dict = apk_app_string_dict.get(str_id, None)                        
                        results_dict[str_id] = {"result": info["result"], "reason": info["reason"] + " (legacy)", "reviewed": True}
                        
        for str_id in apk_app_string_dict.keys():
            if include_previous and str_id in results_dict.keys():
                #print "Result migrated from previous run"
                continue
            if not spec_app_string_dict.get(str_id, False):
                #print "Failed: " + str_id + " missing from spec"
                if len(apk_app_string_dict[str_id].keys()) > 1:
                    results_dict[str_id] ={"result": "failed", "reason": "[3][failed]Has translation but not in spec"}
                    continue
                #print apk_app_string_dict[str_id]
                results_dict[str_id]={"result": "unknown", "reason": "[1][unknown]Missing in spec"}
                continue

            apk_lang_list = apk_app_string_dict[str_id].keys()
            spec_lang_list = spec_app_string_dict[str_id].keys()

            if set(apk_lang_list) != set(spec_lang_list):
                if len(apk_lang_list) > len(spec_lang_list):
                    #print "Failed: " + str_id + " missing lang in spec: " + str([item for item in apk_lang_list if item not in spec_lang_list])
                    results_dict[str_id] = {"result": "failed", "reason": "[2][failed]Missing lang in spec: " + str(set(apk_lang_list)-set(spec_lang_list))}
                    continue
                elif len(apk_lang_list) < len(spec_lang_list):
                    results_dict[str_id] = {"result": "failed", "reason": "[2][failed]Missing lang in apk: " + str(set(spec_lang_list)-set(apk_lang_list))}
                    continue
                else:
                    #If they are the same length but the content are different
                    results_dict[str_id] = {"result": "failed", "reason": "[2][failed]Missing lang in apk: " + str(set(spec_lang_list)-set(apk_lang_list)) + " and Missing lang in spec: " + str(set(apk_lang_list)-set(spec_lang_list))}
                    continue

            if spec_app_string_dict[str_id] != apk_app_string_dict[str_id]:
                mismatch_lang = []
                for lang in apk_app_string_dict[str_id].keys():
                    if apk_app_string_dict[str_id][lang] != spec_app_string_dict[str_id][lang]:
                        mismatch_lang.append(lang)

                results_dict[str_id] = {"result": "failed", "reason": "[1][failed]Mismatch: " + str(mismatch_lang)}
                continue
            else:
                english = apk_app_string_dict[str_id].get("en", None)
                passed_but_empty = []
                passed_but_same = []
                for lang in apk_app_string_dict[str_id].keys():       
                    if apk_app_string_dict[str_id][lang] == "":
                        passed_but_empty.append(lang)
                    elif lang != "en" and english is not None and english == apk_app_string_dict[str_id][lang]:
                        passed_but_same.append(lang)

                if passed_but_empty == apk_app_string_dict[str_id].keys():
                    results_dict[str_id] = {"result": "passed", "reason": "[5][passed]Empty Strings: " + "lang: " + str(passed_but_empty), "reviewed": True}
                elif passed_but_same != []:
                    results_dict[str_id] = {"result": "failed", "reason": "[4][failed]Lang: " + str(passed_but_same) + " is the same as en"}
                else:
                    results_dict[str_id] = {"result": "passed", "reason": "[1][passed]Everything match during current run", "reviewed": True}
        #print "Total APK Key: " + str(len(apk_app_string_dict.keys()))
        #print "Total Spec Key: " + str(len(spec_app_string_dict.keys()))
        #print "Failed: " + str(failed)
        #print "Failed(missing lang): " + str(failed1)
        #print "Failed(Not in spec but has translation): " + str(failed2)
        #print "Passed: " + str(passed)
        #print "Missing Unknown: " + str(unknown)
        #print "Not All translated Unknown: " + str(unknown1)
        return results_dict


if __name__ == "__main__":
    db_info = {"url": "http://hppsrv2.sdg.rd.hpicorp.net:5984/", "user":"service", "password":"service"}
    global_db_connection = CouchWrapper(db_info["url"], "lst_records" , (db_info["user"], db_info["password"]))
    sp = StringProcessor("android", global_db_connection)
    sp.new_run("HP Smart", "/work/PrinterControl-googlestore-debug-8.7.0.73.apk", "/work/what", test_run=True)